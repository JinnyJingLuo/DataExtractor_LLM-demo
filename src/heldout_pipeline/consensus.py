from __future__ import annotations

import math
import re
import string
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl.styles import Font, PatternFill

from .manifest import PaperRecord
from .normalize import normalize_sample_id
from .schema import TABLE1_COLUMNS


DEFAULT_NULL_VALUES = {
    "",
    "-",
    "--",
    "na",
    "n/a",
    "nan",
    "none",
    "null",
    "not reported",
    "not available",
    "unknown",
}

TEXT_FIELDS = {
    "Metal Symbol",
    "Sample ID",
    "Synthesis Method",
    "Treatment",
    "Hardness",
    "Flyer Material Name",
    "Flyer Material Code",
    "Experiment Type",
    "Reference Title",
    "DOI",
    "Verification",
}

DEFAULT_ANCHOR_FIELDS = [
    "Metal Symbol",
    "Initial Temperature (K)",
    "Sample Thickness (mm)",
    "Impact Velocity (m/s)",
    "Flyer Thickness (mm)",
    "Flyer Material Name",
    "Flyer Material Code",
]

STATIC_ANCHOR_FIELDS = {
    "Metal Symbol",
    "Flyer Material Name",
    "Flyer Material Code",
}


@dataclass(frozen=True)
class ConsensusResult:
    consensus_candidates: pd.DataFrame
    disagreement_review: pd.DataFrame
    record_review: pd.DataFrame
    candidate_ground_truth: pd.DataFrame


def _is_null(value: object, null_values: set[str]) -> bool:
    if pd.isna(value):
        return True
    return str(value).strip().casefold() in null_values


def _mechanical_text_key(value: object) -> str:
    if pd.isna(value):
        return ""
    normalized = str(value).strip().casefold()
    normalized = normalized.translate(str.maketrans("", "", string.punctuation))
    return re.sub(r"\s+", "", normalized)


def _numeric_value(value: object) -> float | None:
    if pd.isna(value):
        return None
    try:
        number = float(str(value).strip())
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number):
        return None
    return number


def _is_numeric_field(field_name: str, text_fields: set[str]) -> bool:
    return field_name not in text_fields


def _values_agree(
    field_name: str,
    gemini_value: object,
    claude_value: object,
    *,
    relative_tolerance: float,
    absolute_tolerance: float,
    numeric_floor: float,
    null_values: set[str],
    text_fields: set[str],
) -> tuple[bool, object, str, str]:
    gemini_null = _is_null(gemini_value, null_values)
    claude_null = _is_null(claude_value, null_values)
    if gemini_null and claude_null:
        return True, "", "both_null", ""
    if gemini_null != claude_null:
        return False, "", "", "missing_in_one"

    if _is_numeric_field(field_name, text_fields):
        gemini_number = _numeric_value(gemini_value)
        claude_number = _numeric_value(claude_value)
        if gemini_number is not None and claude_number is not None:
            tolerance = max(
                absolute_tolerance,
                relative_tolerance
                * max(abs(gemini_number), numeric_floor),
            )
            if abs(gemini_number - claude_number) <= tolerance:
                return (
                    True,
                    gemini_number,
                    "numeric_within_tolerance",
                    "",
                )
            return False, "", "", "numeric_mismatch"

    if _mechanical_text_key(gemini_value) == _mechanical_text_key(claude_value):
        return True, gemini_value, "text_mechanical_match", ""
    return False, "", "", "text_mismatch"


def _load_prediction(root: Path, record: PaperRecord) -> pd.DataFrame:
    path = Path(root) / record.split / record.paper_id / "extracted_data.csv"
    if not path.exists():
        return pd.DataFrame(columns=TABLE1_COLUMNS)
    frame = pd.read_csv(path)
    for column in TABLE1_COLUMNS:
        if column not in frame.columns:
            frame[column] = ""
    return frame[TABLE1_COLUMNS].copy()


def _anchor_signature(row: pd.Series, anchor_fields: list[str]) -> tuple[str, ...]:
    parts = []
    dynamic_parts = 0
    for field in anchor_fields:
        value = row.get(field, "")
        if _is_null(value, DEFAULT_NULL_VALUES):
            continue
        number = _numeric_value(value)
        normalized = str(number) if number is not None else _mechanical_text_key(value)
        parts.append(f"{field}={normalized}")
        if field not in STATIC_ANCHOR_FIELDS:
            dynamic_parts += 1
    if len(parts) < 2 or dynamic_parts < 1:
        return ()
    return tuple(parts)


def _unique_index_by(values: Iterable[object]) -> dict[str, int]:
    index: dict[str, int] = {}
    duplicates: set[str] = set()
    for position, value in enumerate(values):
        key = normalize_sample_id(value)
        if not key:
            continue
        if key in index:
            duplicates.add(key)
        else:
            index[key] = position
    for key in duplicates:
        index.pop(key, None)
    return index


def _match_records(
    paper_id: str,
    gemini_frame: pd.DataFrame,
    claude_frame: pd.DataFrame,
    anchor_fields: list[str],
) -> tuple[list[tuple[int, int, str]], list[dict]]:
    matches: list[tuple[int, int, str]] = []
    review_rows: list[dict] = []
    used_gemini: set[int] = set()
    used_claude: set[int] = set()

    gemini_by_id = _unique_index_by(gemini_frame["Sample ID"])
    claude_by_id = _unique_index_by(claude_frame["Sample ID"])
    for sample_key in sorted(set(gemini_by_id) & set(claude_by_id)):
        gemini_index = gemini_by_id[sample_key]
        claude_index = claude_by_id[sample_key]
        matches.append((gemini_index, claude_index, "sample_id"))
        used_gemini.add(gemini_index)
        used_claude.add(claude_index)

    def remaining_signatures(
        frame: pd.DataFrame,
        used: set[int],
    ) -> dict[tuple[str, ...], list[int]]:
        result: dict[tuple[str, ...], list[int]] = {}
        for index, row in frame.iterrows():
            if index in used:
                continue
            signature = _anchor_signature(row, anchor_fields)
            if signature:
                result.setdefault(signature, []).append(index)
        return result

    gemini_by_anchor = remaining_signatures(gemini_frame, used_gemini)
    claude_by_anchor = remaining_signatures(claude_frame, used_claude)
    for signature in sorted(set(gemini_by_anchor) & set(claude_by_anchor)):
        gemini_indexes = gemini_by_anchor[signature]
        claude_indexes = claude_by_anchor[signature]
        if len(gemini_indexes) == 1 and len(claude_indexes) == 1:
            gemini_index = gemini_indexes[0]
            claude_index = claude_indexes[0]
            matches.append((gemini_index, claude_index, "anchor_fields"))
            used_gemini.add(gemini_index)
            used_claude.add(claude_index)
        else:
            for gemini_index in gemini_indexes:
                review_rows.append(
                    {
                        "paper_id": paper_id,
                        "status": "RECORD_REVIEW_REQUIRED",
                        "reason": "ambiguous_anchor_match",
                        "gemini_sample_id": gemini_frame.loc[gemini_index, "Sample ID"],
                        "claude_sample_id": "",
                        "normalized_sample_id": normalize_sample_id(
                            gemini_frame.loc[gemini_index, "Sample ID"]
                        ),
                    }
                )
            for claude_index in claude_indexes:
                review_rows.append(
                    {
                        "paper_id": paper_id,
                        "status": "RECORD_REVIEW_REQUIRED",
                        "reason": "ambiguous_anchor_match",
                        "gemini_sample_id": "",
                        "claude_sample_id": claude_frame.loc[claude_index, "Sample ID"],
                        "normalized_sample_id": normalize_sample_id(
                            claude_frame.loc[claude_index, "Sample ID"]
                        ),
                    }
                )

    for index, row in gemini_frame.iterrows():
        if index not in used_gemini:
            review_rows.append(
                {
                    "paper_id": paper_id,
                    "status": "RECORD_REVIEW_REQUIRED",
                    "reason": "gemini_record_unmatched",
                    "gemini_sample_id": row["Sample ID"],
                    "claude_sample_id": "",
                    "normalized_sample_id": normalize_sample_id(row["Sample ID"]),
                }
            )
    for index, row in claude_frame.iterrows():
        if index not in used_claude:
            review_rows.append(
                {
                    "paper_id": paper_id,
                    "status": "RECORD_REVIEW_REQUIRED",
                    "reason": "claude_record_unmatched",
                    "gemini_sample_id": "",
                    "claude_sample_id": row["Sample ID"],
                    "normalized_sample_id": normalize_sample_id(row["Sample ID"]),
                }
            )
    return matches, review_rows


def _empty_frame(columns: list[str]) -> pd.DataFrame:
    return pd.DataFrame(columns=columns)


def compare_artifact_roots(
    gemini_artifact_root: Path,
    claude_artifact_root: Path,
    records: list[PaperRecord],
    *,
    relative_tolerance: float = 0.005,
    absolute_tolerance: float = 0.0,
    numeric_floor: float = 1e-12,
    null_values: Iterable[str] = DEFAULT_NULL_VALUES,
    categorical_fields: Iterable[str] | None = None,
    anchor_fields: list[str] | None = None,
) -> ConsensusResult:
    null_set = {str(value).strip().casefold() for value in null_values}
    text_fields = set(categorical_fields or set()) | TEXT_FIELDS
    anchors = anchor_fields or DEFAULT_ANCHOR_FIELDS
    candidate_rows: list[dict] = []
    disagreement_rows: list[dict] = []
    record_review_rows: list[dict] = []
    ground_truth_rows: list[dict] = []

    for record in records:
        gemini_frame = _load_prediction(gemini_artifact_root, record)
        claude_frame = _load_prediction(claude_artifact_root, record)
        matches, paper_review_rows = _match_records(
            record.paper_id,
            gemini_frame,
            claude_frame,
            anchors,
        )
        record_review_rows.extend(paper_review_rows)

        for gemini_index, claude_index, match_method in matches:
            gemini_row = gemini_frame.loc[gemini_index]
            claude_row = claude_frame.loc[claude_index]
            normalized_sample = normalize_sample_id(gemini_row["Sample ID"])
            output_row = {
                "sheet": record.paper_id,
                "Sample ID": gemini_row["Sample ID"],
            }
            for field_name in TABLE1_COLUMNS:
                if field_name == "Sample ID":
                    continue
                agree, value, agreement_type, reason = _values_agree(
                    field_name,
                    gemini_row.get(field_name, ""),
                    claude_row.get(field_name, ""),
                    relative_tolerance=relative_tolerance,
                    absolute_tolerance=absolute_tolerance,
                    numeric_floor=numeric_floor,
                    null_values=null_set,
                    text_fields=text_fields,
                )
                if agree:
                    output_row[field_name] = value
                    candidate_rows.append(
                        {
                            "paper_id": record.paper_id,
                            "sample_id": gemini_row["Sample ID"],
                            "claude_sample_id": claude_row["Sample ID"],
                            "normalized_sample_id": normalized_sample,
                            "field_name": field_name,
                            "value": value,
                            "gemini_value": gemini_row.get(field_name, ""),
                            "claude_value": claude_row.get(field_name, ""),
                            "agreement_type": agreement_type,
                            "match_method": match_method,
                        }
                    )
                else:
                    output_row[field_name] = ""
                    disagreement_rows.append(
                        {
                            "paper_id": record.paper_id,
                            "gemini_sample_id": gemini_row["Sample ID"],
                            "claude_sample_id": claude_row["Sample ID"],
                            "normalized_sample_id": normalized_sample,
                            "field_name": field_name,
                            "gemini_value": gemini_row.get(field_name, ""),
                            "claude_value": claude_row.get(field_name, ""),
                            "reason": reason,
                            "match_method": match_method,
                        }
                    )
            ground_truth_rows.append(output_row)

    candidate_columns = [
        "paper_id",
        "sample_id",
        "claude_sample_id",
        "normalized_sample_id",
        "field_name",
        "value",
        "gemini_value",
        "claude_value",
        "agreement_type",
        "match_method",
    ]
    disagreement_columns = [
        "paper_id",
        "gemini_sample_id",
        "claude_sample_id",
        "normalized_sample_id",
        "field_name",
        "gemini_value",
        "claude_value",
        "reason",
        "match_method",
    ]
    review_columns = [
        "paper_id",
        "status",
        "reason",
        "gemini_sample_id",
        "claude_sample_id",
        "normalized_sample_id",
    ]
    gt_columns = ["sheet", "Sample ID"] + [
        column for column in TABLE1_COLUMNS if column != "Sample ID"
    ]
    return ConsensusResult(
        consensus_candidates=(
            pd.DataFrame(candidate_rows, columns=candidate_columns)
            if candidate_rows
            else _empty_frame(candidate_columns)
        ),
        disagreement_review=(
            pd.DataFrame(disagreement_rows, columns=disagreement_columns)
            if disagreement_rows
            else _empty_frame(disagreement_columns)
        ),
        record_review=(
            pd.DataFrame(record_review_rows, columns=review_columns)
            if record_review_rows
            else _empty_frame(review_columns)
        ),
        candidate_ground_truth=(
            pd.DataFrame(ground_truth_rows, columns=gt_columns)
            if ground_truth_rows
            else _empty_frame(gt_columns)
        ),
    )


def _collapse_evidence(evidence: pd.DataFrame, prefix: str) -> pd.DataFrame:
    columns = [
        "paper_id",
        "field_name",
        f"{prefix}_evidence_location",
        f"{prefix}_evidence_notes",
    ]
    if evidence is None or evidence.empty:
        return pd.DataFrame(columns=columns)
    required = {"paper_id", "field_name", "source_location", "notes"}
    missing = required - set(evidence.columns)
    if missing:
        raise ValueError(f"{prefix} evidence missing columns: {sorted(missing)}")
    frame = evidence.copy()
    frame["paper_id"] = frame["paper_id"].astype(str).str.strip()
    frame["field_name"] = frame["field_name"].astype(str).str.strip()

    def join_unique(values: pd.Series) -> str:
        result = []
        for value in values:
            if pd.isna(value):
                continue
            text = str(value).strip()
            if text and text not in result:
                result.append(text)
        return " | ".join(result)

    collapsed = (
        frame.groupby(["paper_id", "field_name"], dropna=False)
        .agg(
            source_location=("source_location", join_unique),
            notes=("notes", join_unique),
        )
        .reset_index()
        .rename(
            columns={
                "source_location": f"{prefix}_evidence_location",
                "notes": f"{prefix}_evidence_notes",
            }
        )
    )
    return collapsed[columns]


def _with_evidence(
    frame: pd.DataFrame,
    gemini_evidence: pd.DataFrame | None,
    claude_evidence: pd.DataFrame | None,
) -> pd.DataFrame:
    result = frame.copy()
    for evidence, prefix in [
        (gemini_evidence, "gemini"),
        (claude_evidence, "claude"),
    ]:
        collapsed = _collapse_evidence(evidence, prefix)
        result = result.merge(
            collapsed,
            on=["paper_id", "field_name"],
            how="left",
        )
    return result.fillna("")


def _style_manual_review(path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="FF1F4E78")
    review_fill = PatternFill("solid", fgColor="FFFFF2CC")
    record_fill = PatternFill("solid", fgColor="FFFCE4D6")
    black_font = Font(color="FF000000")
    header_font = Font(color="FFFFFFFF", bold=True)

    for sheet in workbook.worksheets:
        if sheet.max_row < 1:
            continue
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = black_font
            if sheet.title == "manual_field_review":
                for cell in row:
                    cell.fill = review_fill
            elif sheet.title == "record_review":
                for cell in row:
                    cell.fill = record_fill
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            adjusted = min(max(max_length + 2, 12), 60)
            sheet.column_dimensions[column_cells[0].column_letter].width = adjusted
    workbook.save(path)


def _write_manual_review_workbook(
    result: ConsensusResult,
    output_dir: Path,
    *,
    gemini_evidence: pd.DataFrame | None = None,
    claude_evidence: pd.DataFrame | None = None,
) -> None:
    manual = _with_evidence(
        result.disagreement_review,
        gemini_evidence,
        claude_evidence,
    ).rename(
        columns={
            "gemini_value": "Gemini generated result",
            "claude_value": "Claude Opus generated result",
            "gemini_evidence_location": "Gemini evidence location",
            "gemini_evidence_notes": "Gemini evidence notes",
            "claude_evidence_location": "Claude Opus evidence location",
            "claude_evidence_notes": "Claude Opus evidence notes",
        }
    )
    if not manual.empty:
        manual.insert(0, "manual_status", "")
        manual.insert(1, "adjudicated_value", "")
        manual.insert(2, "reviewer_notes", "")

    agreed = _with_evidence(
        result.consensus_candidates,
        gemini_evidence,
        claude_evidence,
    ).rename(
        columns={
            "gemini_value": "Gemini generated result",
            "claude_value": "Claude Opus generated result",
            "gemini_evidence_location": "Gemini evidence location",
            "gemini_evidence_notes": "Gemini evidence notes",
            "claude_evidence_location": "Claude Opus evidence location",
            "claude_evidence_notes": "Claude Opus evidence notes",
        }
    )

    if manual.empty:
        by_paper = pd.DataFrame(columns=["paper_id", "manual_review_rows"])
        by_field = pd.DataFrame(columns=["field_name", "manual_review_rows"])
    else:
        by_paper = (
            manual["paper_id"]
            .value_counts()
            .rename_axis("paper_id")
            .reset_index(name="manual_review_rows")
        )
        by_field = (
            manual["field_name"]
            .value_counts()
            .rename_axis("field_name")
            .reset_index(name="manual_review_rows")
        )
    summary = pd.DataFrame(
        [
            {"metric": "agreed_fields", "count": len(result.consensus_candidates)},
            {"metric": "field_disagreements", "count": len(result.disagreement_review)},
            {"metric": "record_review_rows", "count": len(result.record_review)},
        ]
    )

    workbook_path = output_dir / "manual_review_with_evidence.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="summary", index=False)
        by_paper.to_excel(writer, sheet_name="review_by_paper", index=False)
        by_field.to_excel(writer, sheet_name="review_by_field", index=False)
        manual.to_excel(writer, sheet_name="manual_field_review", index=False)
        result.record_review.to_excel(writer, sheet_name="record_review", index=False)
        agreed.to_excel(writer, sheet_name="agreed_fields", index=False)
        result.candidate_ground_truth.to_excel(
            writer,
            sheet_name="candidate_ground_truth",
            index=False,
        )
    _style_manual_review(workbook_path)


def write_consensus_outputs(
    result: ConsensusResult,
    output_dir: Path,
    *,
    gemini_evidence: pd.DataFrame | None = None,
    claude_evidence: pd.DataFrame | None = None,
) -> None:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    result.consensus_candidates.to_csv(
        output_dir / "consensus_candidates.csv",
        index=False,
    )
    result.disagreement_review.to_csv(
        output_dir / "disagreement_review.csv",
        index=False,
    )
    result.record_review.to_csv(output_dir / "record_review.csv", index=False)
    result.candidate_ground_truth.to_csv(
        output_dir / "candidate_ground_truth.csv",
        index=False,
    )
    with pd.ExcelWriter(
        output_dir / "candidate_ground_truth.xlsx",
        engine="openpyxl",
    ) as writer:
        result.candidate_ground_truth.to_excel(writer, index=False, startrow=1)
    _write_manual_review_workbook(
        result,
        output_dir,
        gemini_evidence=gemini_evidence,
        claude_evidence=claude_evidence,
    )
